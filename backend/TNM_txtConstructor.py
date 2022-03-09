#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import time , os
deskname=os.getlogin()
print("Hi! Welcome to TNMtxt_constructor by Sajjad Rezvani.K ... \n ")

fileName_exel= input('Enter your exel file name on Desktop:   ')
fileName1='C:\\Users\\'+ deskname +'\\Desktop\\'+ fileName_exel

fileName_sheet= input('Enter your exel sheet name:   ')

print("watting...\n       ....\n       .....\n\n")

os.mkdir('C:\\Users\\'+deskname+'\\Desktop\\TNMtxt_constructor')
fileName_txt='C:\\Users\\'+deskname+'\\Desktop\\TNMtxt_constructor'

#myfile= load_workbook('./me17_DTC.xlsx')
myfile= load_workbook(fileName1)
sheet = myfile.get_sheet_by_name(fileName_sheet)

nameC = input('\nEnter file name collumn in capital:   ')
contC = input('Enter file contents collumn in capital:   ')

firstR = int(input("\nEnter number of first row:   "))
lastR = int(input("Enter number of last row:  "))


i =0
for i in range(firstR,lastR+1):
    name= nameC+ str(i)
    content= contC+ str(i)
    fileName3= fileName_txt+ '\\' +sheet[name].value + '.txt'
    file1= open(fileName3,'w',encoding="utf-8")
    if(sheet[content].value!=None):
        file1.write(sheet[content].value)
    file1.close()

print("\n \n \n good luck! Enjoy of it... ")
`time.sleep(7)